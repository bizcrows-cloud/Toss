# -*- coding: utf-8 -*-
"""
app.py (Streamlit 전용)
- GitHub → Streamlit Cloud 배포를 가정한 단일 파일
- 불필요한 Flask/자동설치 제거
- 결과 파일은 /tmp(또는 OS 임시 디렉토리)에 저장 후 download_button 제공
"""
import os, re, math
from datetime import datetime
import tempfile
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ----------------------------
# 저장 경로: 스트림릿/클라우드에서 안전한 임시 디렉토리 사용
# ----------------------------
OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "toss_cash_outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ----------------------------
# 스타일/파라미터
# ----------------------------
FILL_YELLOW = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")  # 노랑
FILL_RED    = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")  # 빨강
FILL_SKY    = PatternFill(start_color="00ADD8E6", end_color="00ADD8E6", fill_type="solid")  # 하늘색

MATCH_TIME_WINDOWS = [i for i in range(30, 301, 30)]  # 승인번호 매칭: 정확(0) 후 ±30~±300초
POST_TIME_WINDOWS  = [0, 30, 60, 120]                 # E단계: 0, ±30, ±60, ±120
PRESERVE_LEADING_ZERO_IN_H = False                    # H열 선행 0 유지 여부

# ----------------------------
# 유틸
# ----------------------------
def coerce_number(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace(",", "").replace("₩","").strip()
    try: return float(s)
    except ValueError: return None

def coerce_datetime(val):
    from datetime import datetime as _dt
    if isinstance(val, _dt): return val
    if val is None: return None
    s = str(val).strip()
    if not s: return None
    fmts = [
        "%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S","%Y/%m/%d %H:%M",
        "%Y.%m.%d %H:%M:%S","%Y.%m.%d %H:%M",
        "%Y-%m-%d","%Y/%m/%d","%Y.%m.%d",
    ]
    for f in fmts:
        try: return _dt.strptime(s, f)
        except ValueError: pass
    m = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})(?:[ T](\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?", s)
    if m:
        y, mo, d, hh, mm, ss = m.groups()
        hh = int(hh) if hh else 0
        mm = int(mm) if mm else 0
        ss = int(ss) if ss else 0
        try:
            return _dt(int(y), int(mo), int(d), hh, mm, ss)
        except ValueError:
            pass
    m = re.search(r"(\d{14})", s)  # 20250102123456
    if m:
        try:
            return _dt.strptime(m.group(1), "%Y%m%d%H%M%S")
        except ValueError:
            pass
    return None

def dt_fmt(dt): 
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def extract_dt_from_tid(tid_value):
    """V열 TID에서 숫자 시작 이후 연속 숫자 중 앞 14자리 → datetime(YYYYMMDDhhmmss)"""
    if tid_value is None: return None
    s = str(tid_value)
    m = re.search(r"(\d+)", s)
    if not m: return None
    digits = m.group(1)
    if len(digits) < 14: return None
    try:
        from datetime import datetime as _dt
        return _dt.strptime(digits[:14], "%Y%m%d%H%M%S")
    except ValueError:
        return None

def date_only(dt):
    return dt.date() if hasattr(dt, "date") else None

def is_red_fill(cell):
    if cell.fill is None or cell.fill.fill_type != "solid": return False
    rgb = (getattr(cell.fill.start_color, "rgb", None) or "").upper()
    return rgb in {"00FF0000", "FFFF0000"}

def copy_row_values(src_ws, src_row, dst_ws, dst_row, max_col):
    for c in range(1, max_col+1):
        dst_ws.cell(row=dst_row, column=c, value=src_ws.cell(row=src_row, column=c).value)

def delete_rows_desc(ws, rows):
    for r in sorted(rows, reverse=True):
        ws.delete_rows(r, 1)

def ensure_third_sheet(wb, title="전처리_이동"):
    if len(wb.worksheets) >= 3:
        ws3 = wb.worksheets[2]
        ws3.title = title
        for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
            for c in row:
                c.value = None
                c.fill = None
        return ws3
    else:
        return wb.create_sheet(title)

def norm_h(value):
    """H열 키 정규화: 숫자만 남기고(하이픈/공백 제거). 선행 0 유지 옵션."""
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\D+", "", s)
    if not s: return ""
    if PRESERVE_LEADING_ZERO_IN_H:
        return s
    return s.lstrip("0") or "0"

def final_recheck_cash_g(ws_toss, ws_cash, red_fill):
    """최종 상태 기준으로 toss.T/U와 중복되지 않는 cash.G만 빨간색으로 음영."""
    from openpyxl.styles import PatternFill as _PF
    used = set()
    for r in range(2, ws_toss.max_row + 1):
        v_t = ws_toss.cell(row=r, column=20).value  # T
        v_u = ws_toss.cell(row=r, column=21).value  # U
        if v_t not in (None, ""): used.add(str(v_t))
        if v_u not in (None, ""): used.add(str(v_u))
    for r in range(2, ws_cash.max_row + 1):
        g_cell = ws_cash.cell(row=r, column=7)
        # 기존 색 제거
        if g_cell.fill and g_cell.fill.fill_type == "solid":
            g_cell.fill = _PF(fill_type=None)
        v = g_cell.value
        if v in (None, ""): 
            continue
        if str(v) not in used:
            g_cell.fill = red_fill

# ----------------------------
# 메인 처리
# ----------------------------
def process_workbook(input_path: str, output_dir: str) -> str:
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"파일이 없습니다: {input_path}")

    wb = load_workbook(input_path)
    if len(wb.worksheets) < 2:
        raise RuntimeError("시트가 2개 미만입니다. (1:toss, 2:cash)")

    ws_toss = wb.worksheets[0]  # 1: toss
    ws_cash = wb.worksheets[1]  # 2: cash
    ws_move = ensure_third_sheet(wb, title="전처리_이동")

    # 오늘 날짜 파일명
    today_str = datetime.now().strftime("%Y-%m-%d")
    output_filename = f"cash_{today_str}.xlsx"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, output_filename)

    # ----- 전처리 A: toss 그룹 상쇄 이동 -----
    max_col = max(ws_toss.max_column, ws_cash.max_column)
    ws_move.cell(row=1, column=1, value="source")

    groups_toss = {}
    for r in range(2, ws_toss.max_row + 1):
        order_no = ws_toss.cell(row=r, column=6).value  # F
        amt_l = coerce_number(ws_toss.cell(row=r, column=12).value)  # L
        if order_no is None or amt_l is None: continue
        groups_toss.setdefault(order_no, []).append((r, amt_l))

    rows_to_move_toss = set()
    for key, rows in groups_toss.items():
        pos_sum = sum(a for (_, a) in rows if a > 0)
        neg_sum = sum(a for (_, a) in rows if a < 0)
        if pos_sum > 0 and abs(neg_sum) > 0 and math.isclose(pos_sum, abs(neg_sum), rel_tol=0, abs_tol=0.5):
            for (r, _) in rows: rows_to_move_toss.add(r)

    move_cursor = 2
    for r in sorted(rows_to_move_toss):
        ws_move.cell(row=move_cursor, column=1, value="toss")
        copy_row_values(ws_toss, r, ws_move, move_cursor, max_col)
        move_cursor += 1
    delete_rows_desc(ws_toss, rows_to_move_toss)

    # ----- 전처리 B: cash 그룹 상쇄 이동 (같은 '날짜' + 같은 H 정규화) -----
    groups_cash = {}
    for r in range(2, ws_cash.max_row + 1):
        dt_b = coerce_datetime(ws_cash.cell(row=r, column=2).value)  # B
        key_date = date_only(dt_b)
        if key_date is None:
            raw = str(ws_cash.cell(row=r, column=2).value).strip()
            m = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", raw)
            if m:
                y, mo, d = map(int, m.groups())
                try:
                    key_date = datetime(y, mo, d).date()
                except ValueError:
                    key_date = None

        h_raw = ws_cash.cell(row=r, column=8).value
        h_key = norm_h(h_raw)

        amt_f = coerce_number(ws_cash.cell(row=r, column=6).value)

        if key_date is None or not h_key or amt_f is None:
            continue
        groups_cash.setdefault((key_date, h_key), []).append((r, amt_f))

    rows_to_move_cash = set()
    for key, rows in groups_cash.items():
        pos_sum = sum(a for (_, a) in rows if a > 0)
        neg_sum = sum(a for (_, a) in rows if a < 0)
        if pos_sum > 0 and abs(neg_sum) > 0 and math.isclose(pos_sum, abs(neg_sum), rel_tol=0, abs_tol=0.5):
            for (r, _) in rows: rows_to_move_cash.add(r)

    for r in sorted(rows_to_move_cash):
        ws_move.cell(row=move_cursor, column=1, value="cash")
        copy_row_values(ws_cash, r, ws_move, move_cursor, max_col)
        move_cursor += 1
    delete_rows_desc(ws_cash, rows_to_move_cash)

    # ----- 본처리 C: 승인번호 매칭 -----
    targets = []
    for r in range(2, ws_toss.max_row + 1):
        if ws_toss.cell(row=r, column=21).value not in (None, ""):  # U
            continue
        tid_val = ws_toss.cell(row=r, column=22).value  # V
        dt_from_tid = extract_dt_from_tid(tid_val)
        if dt_from_tid is None: continue
        ws_toss.cell(row=r, column=3, value=dt_fmt(dt_from_tid))  # C
        amt_l = coerce_number(ws_toss.cell(row=r, column=12).value)  # L
        if amt_l is None: continue
        targets.append((r, dt_from_tid, amt_l))

    by_amount = {}
    for r in range(2, ws_cash.max_row + 1):
        dt_b = coerce_datetime(ws_cash.cell(row=r, column=2).value)  # B
        amt_f = coerce_number(ws_cash.cell(row=r, column=6).value)   # F
        appr_g = ws_cash.cell(row=r, column=7).value                 # G
        if dt_b is None or amt_f is None or appr_g in (None, ""): continue
        appr_g = str(appr_g)
        by_amount.setdefault(amt_f, []).append((dt_b, appr_g))

    used_approvals = set()
    for r in range(2, ws_toss.max_row + 1):
        uval = ws_toss.cell(row=r, column=21).value
        if uval not in (None, ""):
            used_approvals.add(str(uval))

    for (toss_row, toss_dt, toss_amt) in targets:
        appr = None
        # 정확(0)
        for (c_dt, c_appr) in by_amount.get(toss_amt, []):
            if c_appr in used_approvals: continue
            if c_dt == toss_dt:
                appr = c_appr
                break
        # 단계 확대
        if not appr:
            for w in MATCH_TIME_WINDOWS:
                best, best_diff = None, None
                for (c_dt, c_appr) in by_amount.get(toss_amt, []):
                    if c_appr in used_approvals: continue
                    diff = abs((c_dt - toss_dt).total_seconds())
                    if diff <= w:
                        if best_diff is None or diff < best_diff:
                            best_diff, best = diff, c_appr
                if best:
                    appr = best
                    break
        if appr:
            t_cell = ws_toss.cell(row=toss_row, column=20)  # T
            t_cell.value = appr
            t_cell.fill = FILL_YELLOW
            used_approvals.add(appr)

    # ----- 중간: toss.T/U에 없는 cash.G 빨간색 표시 -----
    def shade_cash_g_not_in_toss():
        used = set()
        for r in range(2, ws_toss.max_row + 1):
            for col in (20, 21):  # T, U
                v = ws_toss.cell(row=r, column=col).value
                if v not in (None, ""):
                    used.add(str(v))
        for r in range(2, ws_cash.max_row + 1):
            g_cell = ws_cash.cell(row=r, column=7)
            v = g_cell.value
            if v in (None, ""): continue
            if str(v) not in used:
                g_cell.fill = FILL_RED
    shade_cash_g_not_in_toss()

    # ----- E단계: T/U 비어있는 toss행 ↔ '빨간 cash 행' 0/30/60/120초 매칭 -----
    red_cash_rows = []
    for r in range(2, ws_cash.max_row + 1):
        g_cell = ws_cash.cell(row=r, column=7)
        if g_cell.value in (None, ""): continue
        if is_red_fill(g_cell):
            dt_b = coerce_datetime(ws_cash.cell(row=r, column=2).value)  # B
            amt_f = coerce_number(ws_cash.cell(row=r, column=6).value)   # F
            appr_g = str(ws_cash.cell(row=r, column=7).value)
            if dt_b is None or amt_f is None: continue
            red_cash_rows.append({"row": r, "dt": dt_b, "amt": amt_f, "appr": appr_g, "used": False})

    toss_candidates = []
    for r in range(2, ws_toss.max_row + 1):
        if ws_toss.cell(row=r, column=20).value not in (None, ""):  # T
            continue
        if ws_toss.cell(row=r, column=21).value not in (None, ""):  # U
            continue
        dt_c = coerce_datetime(ws_toss.cell(row=r, column=3).value)  # C
        if dt_c is None: continue
        toss_candidates.append((r, dt_c))
    toss_candidates.sort(key=lambda x: x[0], reverse=True)

    for toss_row, toss_dt in toss_candidates:
        matched_idx = None
        matched_cash = None
        for w in POST_TIME_WINDOWS:
            best, best_diff, best_idx = None, None, None
            for i, rc in enumerate(red_cash_rows):
                if rc["used"]: continue
                diff = abs((rc["dt"] - toss_dt).total_seconds())
                if (w == 0 and diff == 0) or (w > 0 and diff <= w):
                    if best_diff is None or diff < best_diff:
                        best_diff, best, best_idx = diff, rc, i
            if best is not None:
                matched_idx, matched_cash = best_idx, best
                break
        if matched_cash is None:
            continue

        orig_L = coerce_number(ws_toss.cell(row=toss_row, column=12).value)
        insert_row = toss_row + 1
        ws_toss.insert_rows(insert_row, 1)
        ws_toss.cell(row=insert_row, column=3,  value=dt_fmt(matched_cash["dt"]))   # C
        l_ins_cell = ws_toss.cell(row=insert_row, column=12, value=matched_cash["amt"])  # L
        l_ins_cell.fill = FILL_SKY
        ws_toss.cell(row=insert_row, column=20, value=matched_cash["appr"])         # T

        if orig_L is not None and matched_cash["amt"] is not None:
            a, b = abs(float(orig_L)), abs(float(matched_cash["amt"]))
            ws_toss.cell(row=toss_row, column=12, value=(max(a, b) - min(a, b)))
            ws_toss.cell(row=toss_row, column=12).fill = FILL_SKY
            l_ins_cell.fill = FILL_SKY

        red_cash_rows[matched_idx]["used"] = True

    # 최종 재검사 + Q열
    final_recheck_cash_g(ws_toss, ws_cash, FILL_RED)
    for r in range(2, ws_toss.max_row + 1):
        val_t = ws_toss.cell(row=r, column=20).value  # T
        q_val = "현금" if val_t not in (None, "") else "카드"
        ws_toss.cell(row=r, column=17, value=q_val)   # Q열(17)

    wb.save(output_path)
    return output_path

# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="엑셀 처리기", page_icon="📄", layout="centered")
st.title("📄 엑셀 처리기 (Streamlit)")
st.caption("Toss / Cash 2개 시트를 포함한 엑셀(.xlsx)을 업로드하면 규칙에 맞춰 처리 후 결과 파일을 제공합니다.")

uploaded = st.file_uploader("엑셀 파일(.xlsx)을 업로드하세요", type=["xlsx"])

if uploaded is not None:
    # 업로드를 임시 저장
    in_path = os.path.join(OUTPUT_DIR, f"_input_{uploaded.name}")
    with open(in_path, "wb") as f:
        f.write(uploaded.getbuffer())

    if st.button("처리 실행"):
        with st.spinner("처리 중..."):
            try:
                out_path = process_workbook(in_path, OUTPUT_DIR)
                out_name = os.path.basename(out_path)
                with open(out_path, "rb") as f:
                    st.success("완료! 아래 버튼으로 다운로드하세요.")
                    st.download_button(
                        label=f"결과 다운로드: {out_name}",
                        data=f.read(),
                        file_name=out_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"오류: {e}")
